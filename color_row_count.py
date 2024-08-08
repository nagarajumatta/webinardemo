from openpyxl import load_workbook

# Load the workbook
book = load_workbook('E:\Python\output.xlsx')
sheet = book.active

# Define the color code
color_code = 'FF0000'

# Initialize the row count
row_count = 0

# Iterate over the rows in the sheet
for row in sheet.iter_rows():
    # Check if any cell in the row has the specified fill color
    if any(cell.fill.start_color.rgb == color_code for cell in row):
        row_count += 1

print(f'Number of rows filled with color {color_code}: {row_count}')